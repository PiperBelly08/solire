from django.shortcuts import render
from django.views.decorators.http import require_http_methods
from django.http import JsonResponse, HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font

from .helpers.fuzzy_logic import PlantRecommendationFuzzySystem
from .models import SoilCondition
import json

# Initialize the fuzzy system globally or as a singleton
# This avoids re-initializing the system on every request, which can be slow.
fuzzy_system_instance = PlantRecommendationFuzzySystem()

def index(request):
    return render(
        request,
        'solire_app/index.html',
        {
            'soil_conditions': SoilCondition.objects.all(),
        },
    )

@require_http_methods(["POST"])
def insert_data(request):
    try:
        # Parse JSON data from request body
        data = json.loads(request.body)
        
        # Validate pH value before creating the object
        ph_value = float(data.get('ph_value', 0))
        if not (0.0 <= ph_value <= 14.0):
            return JsonResponse({
                'success': False,
                'error': 'pH value must be between 0 and 14',
                'received_ph': ph_value
            }, status=400)

        # Create new instance
        obj = SoilCondition.objects.create(
            temperature_value=data['temperature_c'],
            moisture_value=data['moisture_percent'],
            ph_value=ph_value,
        )

        return JsonResponse({
            'success': True,
            'message': 'Data inserted successfully',
            'id': obj.id,
            'saved_ph': ph_value
        }, status=201)

    except json.JSONDecodeError:
        return JsonResponse({
            'success': False,
            'error': 'Invalid JSON data'
        }, status=400)
    except KeyError as e:
        return JsonResponse({
            'success': False,
            'error': f'Missing required field: {str(e)}',
            'required_fields': ['temperature_c', 'moisture_percent', 'ph_value']
        }, status=400)
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e),
            'type': type(e).__name__
        }, status=500)

@require_http_methods(["GET"])
def list_data(request):
    try:
        soil_conditions = SoilCondition.objects.all().order_by('-timestamps')
        return JsonResponse({
            'success': True,
            'data': list(soil_conditions.values())
        }, status=200)
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)

@require_http_methods(["GET"])
def list_data_with_recommendation(request):
    try:
        soil_conditions = SoilCondition.objects.all().order_by('-timestamps')
        result_data = []

        for sc in soil_conditions:
            recommended_plants_str = ""

            invalid_reasons = []
            if not (0.0 <= sc.ph_value <= 14.0):
                invalid_reasons.append("pH out of range (0-14)")
            if not (0 <= sc.temperature_value <= 50):
                invalid_reasons.append("Temperature out of range (0-50°C)")
            if not (0 <= sc.moisture_value <= 100):
                invalid_reasons.append("Moisture out of range (0-100%)")
            if sc.moisture_value == 0:
                invalid_reasons.append("Moisture is 0%")
            if sc.ph_value == 0:
                invalid_reasons.append("pH is 0")
            if sc.temperature_value == 0:
                invalid_reasons.append("Temperature is 0°C")

            if invalid_reasons:
                recommended_plants_str = f"Invalid Input: {'; '.join(invalid_reasons)}. No recommendation."
            else:
                try:
                    recommended_plants = fuzzy_system_instance.get_plant_recommendation(
                        float(sc.ph_value),
                        sc.temperature_value,
                        sc.moisture_value
                    )

                    if recommended_plants and recommended_plants['all_plants']:
                        recommended_plants_str = ', '.join([
                            f"{plant['plant']} [{plant['suitability_score']:.2f}, {plant['confidence']}]({plant['status']})"
                            for plant in recommended_plants['all_plants']
                        ])
                        if all(plant['suitability_score'] < 0.1 for plant in recommended_plants['all_plants']):
                            recommended_plants_str += " (Note: All plants show very low suitability.)"
                    else:
                        recommended_plants_str = "N/A - No plant recommendations found (possibly due to rule non-firing)."

                except ValueError as ve:
                    recommended_plants_str = f"Fuzzy Logic Input Error: {str(ve)}"
                except Exception as e:
                    recommended_plants_str = f"Error in recommendation logic: {str(e)}"

            # Append result including fuzzy recommendation
            result_data.append({
                'id': sc.id,
                'temperature_value': sc.temperature_value,
                'moisture_value': sc.moisture_value,
                'ph_value': float(sc.ph_value),
                'recommended_plants': recommended_plants_str,
                'timestamps': sc.timestamps.strftime('%d-%m-%Y %H:%M:%S'),
            })

        return JsonResponse({
            'success': True,
            'data': result_data
        }, status=200)

    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e),
            'type': type(e).__name__
        }, status=500)


def clear_data(request):
    try:
        SoilCondition.objects.all().delete()
        return JsonResponse({
            'success': True,
            'message': 'Data cleared successfully'
        }, status=200)
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)


def generate_report(request):
    """
    Generate an Excel report from the database.
    """
    try:
        soil_conditions = SoilCondition.objects.all().order_by('-timestamps')
        wb = Workbook()
        ws = wb.active

        # Set header row
        header_row = ['#', 'Temperature (C)', 'Moisture (%)', 'pH', 'Recommended Plants', 'Timestamps']
        for col, val in enumerate(header_row, start=1):
            cell = ws.cell(row=1, column=col)
            cell.value = val
            cell.font = Font(bold=True)

        # Set data rows
        for row_idx, sc in enumerate(soil_conditions, start=2):
            ws.cell(row=row_idx, column=1).value = row_idx - 1
            ws.cell(row=row_idx, column=2).value = sc.temperature_value
            ws.cell(row=row_idx, column=3).value = sc.moisture_value
            ws.cell(row=row_idx, column=4).value = float(sc.ph_value)

            recommended_plants_str = ""  # Initialize for safety

            # --- Pre-Fuzzy System Validation and Messaging ---
            invalid_reasons = []
            if not (0.0 <= sc.ph_value <= 14.0):
                invalid_reasons.append("pH out of range (0-14)")
            if not (0 <= sc.temperature_value <= 50):
                invalid_reasons.append("Temperature out of range (0-50°C)")
            if not (0 <= sc.moisture_value <= 100):
                invalid_reasons.append("Moisture out of range (0-100%)")

            # Special case for 0 moisture, as it often means 'no data' or extremely dry
            if sc.moisture_value == 0:
                invalid_reasons.append("Moisture is 0%")

            # Special case for pH of 0, as it often means 'no data' or extremely acidic
            if sc.ph_value == 0:
                invalid_reasons.append("pH is 0")

            # Special case for temperature of 0, as it often means 'no data' or extremely cold
            if sc.temperature_value == 0:
                invalid_reasons.append("Temperature is 0°C")

            if invalid_reasons:
                recommended_plants_str = f"Invalid Input: {'; '.join(invalid_reasons)}. No recommendation."
            else:
                try:
                    # Pass parameters in the correct order: (pH, Temperature, Moisture)
                    recommended_plants = fuzzy_system_instance.get_plant_recommendation(
                        float(sc.ph_value),
                        sc.temperature_value,
                        sc.moisture_value
                    )

                    # Format the output from the fuzzy system
                    if recommended_plants and recommended_plants['all_plants']:
                        recommended_plants_str = ', '.join([
                            f"{plant['plant']} [{plant['suitability_score']:.2f}, {plant['confidence']}]({plant['status']})"
                            for plant in recommended_plants['all_plants']
                        ])
                        # If all suitability scores are very low after processing
                        if all(plant['suitability_score'] < 0.1 for plant in recommended_plants['all_plants']):
                            recommended_plants_str += " (Note: All plants show very low suitability.)"
                    else:
                        recommended_plants_str = "N/A - No plant recommendations found (possibly due to rule non-firing)."

                except ValueError as ve:  # Catch validation errors from fuzzy system's internal checks
                    print(f"ERROR: Fuzzy system input validation failed for row {row_idx - 1}: {str(ve)}")
                    recommended_plants_str = f"Fuzzy Logic Input Error: {str(ve)}"
                except Exception as e:
                    print(f"ERROR: General error in get_plant_recommendation for row {row_idx - 1}: {str(e)}")
                    recommended_plants_str = f"Error in recommendation logic: {str(e)}"

            ws.cell(row=row_idx, column=5).value = recommended_plants_str
            ws.cell(row=row_idx, column=6).value = sc.timestamps.strftime('%d-%m-%Y %H:%M:%S')

        # Prepare the file for download
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="soil_conditions_report.xlsx"'
        wb.save(response)

        return response

    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)

def recommend_plant(request):
    if request.method == 'GET':
        # Get parameters from GET request (e.g., /recommend/?ph=6.5&temp=28&humidity=70)
        try:
            ph_value = float(request.GET.get('ph'))
            temp_value = int(request.GET.get('temp'))
            humidity_value = int(request.GET.get('humidity'))
        except (TypeError, ValueError):
            return JsonResponse({'error': 'Invalid or missing input parameters. Please provide ph, temp, and humidity as numbers.'}, status=400)

        try:
            recommendation_results = fuzzy_system_instance.get_plant_recommendation(
                ph_value, temp_value, humidity_value
            )
            return JsonResponse(recommendation_results)
        except ValueError as e:
            return JsonResponse({'error': str(e)}, status=400)
        except Exception as e:
            # Catch any other unexpected errors from the fuzzy system
            return JsonResponse({'error': f'An unexpected error occurred: {e}'}, status=500)

    elif request.method == 'POST':
        # Get parameters from POST request (e.g., JSON payload)
        import json
        try:
            data = json.loads(request.body)
            ph_value = float(data.get('ph'))
            temp_value = int(data.get('temp'))
            humidity_value = int(data.get('humidity'))
        except (json.JSONDecodeError, TypeError, ValueError):
            return JsonResponse({'error': 'Invalid JSON or missing input parameters. Please provide ph, temp, and humidity as numbers.'}, status=400)

        try:
            recommendation_results = fuzzy_system_instance.get_plant_recommendation(
                ph_value, temp_value, humidity_value
            )
            return JsonResponse(recommendation_results)
        except ValueError as e:
            return JsonResponse({'error': str(e)}, status=400)
        except Exception as e:
            return JsonResponse({'error': f'An unexpected error occurred: {e}'}, status=500)

    return JsonResponse({'error': 'Only GET and POST requests are supported.'}, status=405)

def recommendation_form(request):
    # A simple view to render a form for input
    return render(request, 'recommendation/recommendation_form.html')